# For handeling Excel as a file
import openpyxl
from openpyxl import Workbook

# For handeling Excel in memory
import io
import pandas as pd

# Making requests to exportcomments.com
import requests

# Exportcomments.com API
from exportcomments import ExportComments
from exportcomments.exceptions import PlanQueryLimitError, ExportCommentsException #! Not used bc it did not seem to work but put a try catch around the ex.exports.create if any error occurs

# Time to print total runtime and sleep for rate limiting. pkg is not needed but used to specify exportcomments api version.
import time
import pkg_resources

start_time = time.time()

#* Connect to the API and make excel sheet
ex = ExportComments('<PRIVATE_KEY>')

wb = Workbook()
ws = wb.active

def get_urls_from_excel(path):
    # Load the workbook and select active worksheet.
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active

    urls = []

    # Get the URL's (Assuming they are in the first row).
    for row in sheet.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
        url = row[0]

        if url is not None:
            urls.append(url)

    return urls


def generate_excels(url):
    # Use a while true so it keeps trying if rate limit is hit
    while True:
        response = ex.exports.create(url=url, replies=False, twitterType=None)
        
        if "error_code" in response.body['data'] and response.body['data']['error_code'] == 'CONCURRENCY_RATE_LIMIT':
            wait_time = response.body['data']['seconds_to_wait']+1

            print(f"[WAITING FOR RATE LIMIT] time: ({wait_time} seconds)\n")
            time.sleep(wait_time)
        elif "error_code" in response.body['data']: 
            print(f"[UNKNOWN ERROR] error code: {response.body['data']['error_code']}, URL: {url}\n")
            break
        else:
            break

    # Get guid and download_url so we can find it later
    guid = response.body['data']['guid']
    download_url = response.body['data']['downloadUrl']

    print(f"[GENERATE URL] guid: {guid}")
    
    return [guid, download_url, url]


def get_excel(guid):
    # Check if the export is finished
    response = ex.exports.check(guid=guid)

    # Get the status and return if it finished or if there was an error
    status = response.body['data'][0]['status']
    print(f"[CHECK STATUS] guid: {guid}, status: {status}")

    if status == 'done':
        return True, False
        
    elif status == 'error': 
        print(f'[ERROR] There was a problem generating the excel, this one will be skipped (guid: {guid})')
        return False, True
    
    return False, False


def save_files(download_url, url):
    # Set headers for download (might not be necesary but always useful)
    headers = {
        'Authorization': "48f8073b925fe00eb91f3715344a88589a22a16b16d1afbb122d10ba59acaee277740b7c2df820df32fa31987ebc33eaa99ff57c2c295711ef0ac431",
        'Content-Type': 'application/json',
        'User-Agent': 'python-sdk-{}'.format(pkg_resources.get_distribution('exportcomments').version),
    }

    # Get the excel
    response = requests.get("https://exportcomments.com/"+download_url, headers=headers)

    # Handle the excel if is available
    if response.status_code == 200:

        # Load the excel in memory and skip the first 7 rows (starts at 1 for some reason?)
        excel_data = io.BytesIO(response.content)
        df = pd.read_excel(excel_data, skiprows=6)
        
        # Convert the DataFrame to a list of lists, each representing a row in the Excel file
        data_to_append = df.values.tolist()

        # Replace the H column with the URL and append
        for row_data in data_to_append:
            row_data[7] = url  # Replace the H column value with the URL
            ws.append(row_data)

        wb.save("final_results.xlsx")

    
        print(f"[SUCCESSFULL DOWNLOAD ] File Downloaded: {download_url}")

    else:
        print(f"[FAILED TO DOWNLOAD] Status Code: {response.status_code}")


if __name__ == '__main__':
    # Get the urls
    path = "url_list.xlsx"
    urls = get_urls_from_excel(path)

    info = []
    completed_items = []

    # First make them all generate
    [info.append(generate_excels(url)) for url in urls]

    # Wait until every process has finished
    completed = False
    while not completed:
        to_keep = []
        
        for item in info:
            done, error = get_excel(item[0])

            if done: 
                completed_items.append(item)
            elif not error:
                to_keep.append(item)

        info = to_keep
        
        if info == []: completed = True
        else: 
            print("[WAITING] 20 seconds to check if any processes finished.")
            time.sleep(20)

    # Save all the completed excels
    [save_files(item[1], item[2]) for item in completed_items]

    # Save one last time, just to be sure
    wb.save("final_results.xlsx")

    # End of program, print the time it took to run the code
    end_time = time.time()

    total_runtime = end_time - start_time
    print(f"Total runtime: {total_runtime} seconds")